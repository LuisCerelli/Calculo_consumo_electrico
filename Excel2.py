from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Cargar el archivo previamente creado
wb = load_workbook("/Users/luisalbertocerelli/Desktop/00-Todo/03_Python_practicas/Hacer_excel/data/Control_de_Consumo_Energético.xlsx")
ws = wb.active

# Añadir fórmulas en la columna 'Costo por Unidad ($/unidad)'
for row in range(2, 14):
    ws[f'E{row}'] = f"=IFERROR(D{row}/C{row}, 0)"

# Añadir fórmulas para 'Variación de Consumo (%)' y 'Variación de Costo (%)'
for row in range(3, 14):  # Empezar desde la tercera fila para evitar errores en la primera
    ws[f'F{row}'] = f"=IFERROR((C{row}-C{row-1})/C{row-1}*100, 0)"
    ws[f'G{row}'] = f"=IFERROR((D{row}-D{row-1})/D{row-1}*100, 0)"

# Resaltar encabezados en negrita
for col in range(1, 8):
    cell = ws[f"{get_column_letter(col)}1"]
    cell.font = Font(bold=True)

# Guardar el archivo modificado
output_path_ready = "/Users/luisalbertocerelli/Desktop/00-Todo/03_Python_practicas/Hacer_excel/data/Control_de_Consumo_Energético_Listo.xlsx"
wb.save(output_path_ready)

output_path_ready
